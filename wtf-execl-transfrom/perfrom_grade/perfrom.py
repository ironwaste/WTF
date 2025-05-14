import pandas as pd
import os , stat
import openpyxl as oxl
import pypinyin
import xlwt
# 品势成绩转换
from openpyxl.utils import get_column_letter, column_index_from_string
# import MySQLdb
import numpy as np

# 1 -> 姓名
# 2 -> 单位
# 3 -> 级别
# I -> 表现力      (9)
# J -> 准确度       (10)
# K -> 总分        (11)
# L -> 名次        (12)

# output format
# B1:B3 级别 (row , 3)
# C1    姓名 (row , 1)
# C2    单位 (row , 2)
# C3    str(‘总分（准确度，表现力 ）’) ((row,11) ((row,9),(row,10))

# D .... J

# os.chmod('D:\wtf\wtf-execl-transfrom\source' , stat.S_IXUSR)
# # 导入
# shuju1 = pd.read_excel(r'D:\wtf\wtf-execl-transfrom\source\个品儿童男蓝决赛排名.xlsx',sheet_name='个人品势');
# print('\nshuju1\n',shuju1)


# path = r'D:\wtf\wtf-execl-transfrom\source'
# os.chdir(path)
# name = input()

# 格式设置 居中未解决
# def set_style_alignment() :
#     style = xlwt.Alignment()
#
#     style.horz = xlwt.Alignment.HORZ_CENTER
#     style.vert = xlwt.Alignment.VERT_CENTER
#     # style.wrap =
#     return style

# outworkbook = oxl.load_workbook()
# outSheet = outworkbook.active




def set_style(bold=False) :  # 字体格式修改

    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = '黑体' # Times New Roman
    # font.bold = bold    # 加粗
    font.color_index = 4
    font.height = 20 * 8 # 字体大小

    style.font = font
    return style


# workbook = oxl.load_workbook('D:\wtf\wtf-execl-transfrom\source\品势成绩-源数据.xlsx')
# sheetSource= workbook.active



# def write_sort_execl(): # 对于级别 及 名次进行排序
#     df = pd.read_excel('D:\wtf\wtf-execl-transfrom\source\品势成绩-源数据.xlsx')
#     df.sort_values(by='级别', key=lambda x : (pypinyin.lazy_pinyin(x[2])))
#     df.to_excel('D:\wtf\wtf-execl-transfrom\source\品势成绩-sorted数据.xlsx',sheet_name='sheet')
#     workbook = oxl.load_workbook('D:\wtf\wtf-execl-transfrom\source\品势成绩-sorted数据.xlsx')
#     sheetSource = workbook.active
#     return sheetSource

def write_sort_execl(): # 对于级别 及 名次进行排序
    df = pd.read_excel(r'D:\wtf\wtf-execl-transfrom\perfrom_grade\aa.xlsx')
    df.sort_values(by=['级别','名次'],ascending=[True,True])
    df.to_excel(r'D:\wtf\wtf-execl-transfrom\perfrom_grade\aasort.xlsx',sheet_name='sheet')
    workbook = oxl.load_workbook(r'D:\wtf\wtf-execl-transfrom\perfrom_grade\aa.xlsx')
    sheetSource = workbook.active
    return sheetSource


def fromat_trans(Sheet,Row,nameCol,companyNameCol,gradeTecCol,gradePlayCol,gradeSumCol,groupCLassCol) :
    # paramater 1 2

    # 返回格式 ： list[姓名，单位，成绩(‘ 总分 （技术性，表现力）’ )，组别]

    list1 = []
    list1.append(Sheet.cell(Row,nameCol).value)
    list1.append(Sheet.cell(Row,companyNameCol).value)
    list1.append('{}({},{})'.format(Sheet.cell(Row,gradeSumCol).value,Sheet.cell(Row,gradeTecCol).value,Sheet.cell(Row,gradePlayCol).value))
    list1.append(Sheet.cell(Row,groupCLassCol).value)

    return list1




# def write_init() : # 写入初始化
#     return

def write_excel(sheet) :
    # 表头起始位置 以及 合并终止位置
    idStart_row = 0
    idStart_col = 0
    idEnd_row = 1
    idEnd_col = 0

    jibieStart_row = 0
    jibieStart_col = 1
    jibieEnd_row = 1
    jibieEnd_col = 1

    mingciStart_row = 0
    mingciStart_col = 2
    mingciEnd_row = 0
    mingciEnd_col = 9

    # 合并位置定义结束

    outBook = xlwt.Workbook()
    outSheet = outBook.add_sheet(u'sheet1',cell_overwrite_ok=True)

    merge_style = set_style()
    # merge_ali_style = set_style_alignment()
    # outSheet.write_merge(0,3,0,10,'名次',merge_style)


    # tilte -> 标头 列表名称
    outSheet.write_merge(mingciStart_row,mingciEnd_row,mingciStart_col,mingciEnd_col,'名次',merge_style)
    outSheet.write_merge(jibieStart_row,jibieEnd_row,jibieStart_col,jibieEnd_col,'级别',merge_style)
    outSheet.write_merge(idStart_row,idEnd_row,idStart_col,idEnd_col,'序',merge_style)
    col0 = [u'第一名',u'第二名',u'第三名',u'第四名',u'第五名',u'第六名',u'第七名',u'第八名']
    rowName = 1
    for i in range(2,10) :
        outSheet.write(rowName,i,col0[i-2],merge_style)
    # 标头结束

    #初始化 写入位置
    cnt_Id = 1
    cntRow = 1
    cnt_Write_row = 2
    cnt_Write_col = 1
    preZuBie = 'init'
    cntZuBie = 'init'
    rank = 1
    is_Zubie = 0

    MxRow = sheet.max_row + 1

    while cntRow < MxRow :
        preZuBie = cntZuBie
        if sheet.cell(cntRow, 2).value != '姓名' and sheet.cell(cntRow, 2).value != None:
            list = fromat_trans(sheet, cntRow,2,3,10,11,12,4)
            # outSheet.write(,fromat_trans(sheet, cntRow))
            n = len(list)
            cntZuBie = list[n-1]

            # print('cnt_Write: %d   cnt_Write ： %d    rank : %d' % (cnt_Write_row , cnt_Write_col, rank))
            if cntZuBie == preZuBie :
                for i in range(n-1) :
                    outSheet.write(cnt_Write_row + i,cnt_Write_col+rank,list[i])
            elif cntZuBie != preZuBie and preZuBie == 'init' :
                if is_Zubie == 0:
                    outSheet.write_merge(cnt_Write_row, cnt_Write_row + 2, cnt_Write_col, cnt_Write_col, cntZuBie,
                                         merge_style)
                    outSheet.write_merge(cnt_Write_row, cnt_Write_row + 2, cnt_Write_col - 1, cnt_Write_col - 1, cnt_Id,
                                         merge_style)
                    is_Zubie = 1
                    cnt_Id += 1

                for i in range(n - 1):
                    outSheet.write(cnt_Write_row + i, cnt_Write_col + rank, list[i])
            elif cntZuBie != preZuBie :
                rank = 1
                cnt_Write_row += 3
                cnt_Write_col = 1
                for i in range(n - 1):
                    outSheet.write(cnt_Write_row + i, cnt_Write_col + rank, list[i])
                is_Zubie = 0

            if is_Zubie == 0 : # 序号 和 组别合并及写入
                outSheet.write_merge(cnt_Write_row, cnt_Write_row + 2, cnt_Write_col,cnt_Write_col, cntZuBie, merge_style) # 序号
                outSheet.write_merge(cnt_Write_row, cnt_Write_row + 2, cnt_Write_col - 1, cnt_Write_col - 1, cnt_Id,
                                     merge_style)# 组别
                is_Zubie = 1
                cnt_Id += 1

            rank += 1 # 组别内顺序

        cntRow += 1

    outBook.save('./output.xls')

sheet1 = write_sort_execl()

write_excel(sheet1)

