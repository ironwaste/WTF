import pandas as pd
import openpyxl as oxl
import xlwt
import datetime

# 输出格式 ：

# 0、bisaixuhao 	    col ：A  start :                 -- (0)
# 1、zonglunci         empty                           -- (1)
# 2、lunci             format : c(num)
# 3、changdi           format : A B C ....
# 4、chngdiho          format : 2001
# 5、jibie             format : str
# 6、gerentuanti       format : str (个人 / 团体)
# 7、riqi              empty / current data-time
# 8、qingfangbianhao   empty
# 9、qingfangxinming   format ： str (name qing)            9
# 10、qingfangdanwei    format ： str (name danwei )         10
# 11、hongfangbianhao                                       11
# 12、hongfangxinming   format ： str (name hong)            12
# 13、hongfangdanwei    format ： str (name danwei )     --(13)
# 输出 0 -13

# input 输入格式  col 0 - 8
# strat row = 1
# 表名
# 0、场次 str  (char + num) ->  [changdi(4) + chdihao(5)]
# 1、empty
# 2、轮次 -> （map一下 将 final 变成从c1
#              其他带有数字的将非1数字 变成c后面的数字）[2] -> [2]
# 3、青方姓名          [3] -> [9]
# 4、代表队 （单位对应） [4] -> [10]
# 5、 VS
# 6、红方姓名   [6] -> [12]
# 7、代表队    [7] -> [13]
# 8、级别      [8] -> [5]
# 9、备注 bz  empty

def import_sheet():

    workBook = oxl.load_workbook('D:\wtf\wtf-execl-transfrom\source\fight\第一天竞技对阵表.xlsx');
    sheet = workBook.active
    return sheet

def trans_num(lunci) :
    s = str(lunci)
    n = len(s)
    print(s)
    # print(n)
    out ='c'
    for i in range(2,n) :
        out += s[i]
    print(out)
    return out
# 转换 轮次 1/4 为 c4
def trans_changci(changci) :
    s = str(changci)
    return s[0],s[1:5]
# 转换场次号 和 场地
def fromat_trans(Sheet,Row,changCiCol,lunCiCol,qingNameCol,qingCompanyNameCol,hongNameCol,hongCompanyCol,jiBieCol) :
    # 输入 参数，[列表，行，场次（0），轮次（2），青方姓名（3），青单位（4），红姓名（6），红单位（7），级别（8）]

    lunCi = Sheet.cell(Row,lunCiCol).value
    # print('lunci L %s  ' %lunCi)
    lunCiOut = '1'
    if lunCi == 'Final' :
        lunCiOut = 'c1'
    else :
        lunCiOut = trans_num(lunCi)

    changDi,chCiId = trans_changci(Sheet.cell(Row,changCiCol).value)
    list1 = []
    list1.append(lunCiOut)
    list1.append(changDi)
    list1.append(chCiId)
    list1.append(Sheet.cell(Row,jiBieCol).value)
    list1.append('个人')
    list1.append(datetime.datetime.now())
    list1.append(Sheet.cell(Row,qingNameCol).value)
    list1.append(Sheet.cell(Row,qingCompanyNameCol).value)

    list1.append(Sheet.cell(Row,hongNameCol).value)
    list1.append(Sheet.cell(Row,hongCompanyCol).value)
    # 返回格式 ： list[轮次，场地，场次号，级别，个人团体，日期，青方姓名，青方单位，红方姓名，红方单位]
    return list1
# 返回格式 ： list[轮次(2)，场地(3)，场次号(4)，级别5，个人团体6，日期7，青方姓名9，青方单位10，红方姓名12，红方单位13]
# 输入 参数，[列表，行，场次（0），轮次（2），青方姓名（3），青单位（4），红姓名（6），红单位（7），级别（8）]



# NEED : 0(bisaixuhao) 2(lunci) 3(changdi) 4(chngdiho)
# 5(jibie) 6(gerentuanti) 7(riqi) 9(qingfangxinming)
# 10(qingfangdanwei) 12(hongfangxinming) 13(hongfangdanwei)

def write_execl(sheetSource):
    # sheetSource = import_sheet()
    listTitle =['bisaixuhao','zonglunci','lunci','changdi','chngdiho','jibie','gerentuanti','riqi','qingfangbianhao','qingfangxinming','qingfangdanwei','hongfangbianhao','hongfangxinming','hongfangdanwei']
    cntWriteRow = 0

    outBook = xlwt.Workbook()
    outSheet = outBook.add_sheet(u'sheet1',cell_overwrite_ok=True)

    for i in range(len(listTitle) ) :
        outSheet.write(cntWriteRow,i,listTitle[i])
    cntWriteRow += 1
    Mxrow = sheetSource.max_row + 1
    for i in range(2,Mxrow) :

        list1 = fromat_trans(sheetSource,i,1,3,4,5,7,8,9)

        outSheet.write(cntWriteRow,0,cntWriteRow)
        # 2 , 3,4,5,6,7,9,10,12,13
        outSheet.write(cntWriteRow,2,list1[0])
        outSheet.write(cntWriteRow,3,list1[1])
        outSheet.write(cntWriteRow,4,list1[2])
        outSheet.write(cntWriteRow,5,list1[3])
        outSheet.write(cntWriteRow,6,list1[4])
        outSheet.write(cntWriteRow,7,list1[5])

        outSheet.write(cntWriteRow,9,list1[6])
        outSheet.write(cntWriteRow,10,list1[7])
        outSheet.write(cntWriteRow,12,list1[8])
        outSheet.write(cntWriteRow,13,list1[9])
        cntWriteRow += 1

    outBook.save('./output1.xls')



workBook = oxl.load_workbook('D:\wtf\wtf-execl-transfrom\source\\fight\第一天竞技对阵表.xlsx');
sheet = workBook.active


write_execl(sheet)










